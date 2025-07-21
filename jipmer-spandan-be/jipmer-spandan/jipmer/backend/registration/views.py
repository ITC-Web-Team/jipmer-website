from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from django.utils import timezone
from django.http import HttpResponse
from .models import DelegateCardRegistration, EventRegistration, PassPurchase
from .serializers import DelegateCardRegistrationSerializer, EventRegistrationSerializer, PassPurchaseSerializer
import openpyxl
from openpyxl.utils import get_column_letter
from registration.utils.email import send_smtp_email
import json
from django.db import transaction
from rest_framework.parsers import MultiPartParser, FormParser

import logging
logger = logging.getLogger(__name__)

# Constants
EXEMPT_EVENTS = {
    'snap_saga', 'verse', 'reel_realm',
    'junior_mediquiz', 'senior_mediquiz',
    'case_presentation', 'poster_presentation',
    'research_paper', 'mediquiz', 'clinico_pathological_conference',
}
CULT_EVENTS = {'solo_dance', 'duet_dance', 'mono_act', 'street_play', 'meme_making', 'face_painting'}
CULT_MAJOR_EVENTS = {'chorea', 'alaap', 'tinnitus', 'dernier_cri'}
LIT_EVENTS = {'debate', 'poetry', 'essay', 'elocution', 'formal_debate'}
QUIZ_EVENTS = {'medical_quiz', 'general_quiz', 'india_quiz', 'fandom_quiz', 'sports_quiz', 'rewind_quiz', 'tj_jaishankar_memorial_quiz', 'jam'}
BADMINTON_EVENTS = {
    'badminton_men_singles', 'badminton_women_singles',
    'badminton_men_doubles', 'badminton_women_doubles',
    'badminton_mixed_doubles'
}
TENNIS_EVENTS = {
    'tennis_men_singles', 'tennis_women_singles',
    'tennis_men_doubles', 'tennis_women_doubles',
    'tennis_mixed_doubles'
}
TABLE_TENNIS_EVENTS = {
    'table_tennis_men_singles', 'table_tennis_women_singles',
    'table_tennis_men_doubles', 'table_tennis_women_doubles'
}
ATHLETICS_EVENTS = {
    'athletics_100m', 'athletics_200m', 'athletics_400m',
    'athletics_800m', 'athletics_1500m', 'athletics_4x100m',
    'athletics_shot_put', 'athletics_discus_throw', 'athletics_long_jump'
}
AQUATICS_EVENTS = {
    'aquatics_men_50m_freestyle', 'aquatics_men_50m_backstroke',
    'aquatics_men_50m_breaststroke', 'aquatics_men_50m_butterfly',
    'aquatics_men_4x50m_freestyle_relay',
    'aquatics_women_50m_freestyle', 'aquatics_women_50m_backstroke',
    'aquatics_women_50m_breaststroke', 'aquatics_women_50m_butterfly',
    'aquatics_mixed_4x50m_freestyle_relay'
}
SPORTS_EVENTS = {
    'cricket', 'football', 'basketball_men', 'basketball_women',
    'volleyball_men', 'volleyball_women', 'hockey_men',
    'futsal_men', 'futsal_women', 'chess_bullet', 'chess_rapid',
    'chess_blitz', 'chess_team', 'carroms', 'throwball_men', 'throwball_women',
    'badminton', 'tennis', 'table_tennis', 'athletics', 'aquatics'
}

CATEGORY_EVENT_MAP = {
    'fine_arts': {
        'face_painting', 'pot_painting', 'mehendi', 'sketching', 'painting'
    },
    'dance': {
        'solo_dance', 'duet_dance', 'chorea_theme', 'chorea_nontheme', 'show_down'
    },
    'music': {
        'tinnitus', 'alaap', 'euphony', 'raag_rangmanch', 'solo_singing', 'solo_instrumental'
    },
    'drama': {
        'play', 'skit', 'mime', 'adzap', 'variety', 'dernier_cri'
    },
    'sports': {
        'cricket', 'football', 'basketball_men', 'basketball_women', 'volleyball_men',
        'volleyball_women', 'hockey_men', 'futsal_men', 'futsal_women', 'chess_bullet',
        'chess_rapid', 'chess_blitz', 'chess_team', 'carroms', 'throwball_men', 'throwball_women',
        'tennis_men_singles', 'tennis_women_singles', 'tennis_men_doubles','tennis_women_doubles','tennis_mixed_doubles', 'aquatics_men_50m_freestyle', 'aquatics_men_50m_backstroke',
        'aquatics_men_50m_breaststroke', 'aquatics_men_50m_butterfly',
        'aquatics_men_4x50m_freestyle_relay',
        'aquatics_women_50m_freestyle', 'aquatics_women_50m_backstroke',
        'aquatics_women_50m_breaststroke', 'aquatics_women_50m_butterfly',
        'aquatics_mixed_4x50m_freestyle_relay', 'badminton_men_singles', 'badminton_women_singles', 'badminton_men_doubles','badminton_women_doubles','badminton_mixed_doubles', 'table_tennis_men_singles', 'table_tennis_women_singles',
        'table_tennis_men_doubles', 'table_tennis_women_doubles', 'athletics_100m', 'athletics_200m', 'athletics_400m',
        'athletics_800m', 'athletics_1500m', 'athletics_4x100m',
        'athletics_shot_put', 'athletics_discus_throw', 'athletics_long_jump',
    },
    'literary': {
        'malarkey', 'shipwrecked', 'turncoat', 'scrabble', 'formal_debate',
        'cryptic_crossword', 'ppt_karaoke', 'potpourri'
    },
    'quiz': {
        'india_quiz', 'fandom_quiz', 'sports_quiz', 'rewind_quiz',
        'tj_jaishankar_memorial_quiz', 'jam'
    }
}

# Delegate Card Views
class DelegateCardRegisterView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = DelegateCardRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            reg = serializer.save()
            return Response({
                "message": "Registration submitted",
                "registration": DelegateCardRegistrationSerializer(reg).data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        self.check_permissions(request)
        regs = DelegateCardRegistration.objects.all()
        serializer = DelegateCardRegistrationSerializer(regs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def verify_registration(request, pk):
    try:
        reg = DelegateCardRegistration.objects.get(pk=pk)
    except DelegateCardRegistration.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
    
    reg.is_verified = True
    reg.status = 'approved'
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="✅ Delegate Card Verified – Spandan 2025",
        message=f"Dear {reg.name},\n\nWe are delighted to confirm your registration as a delegate for Spandan 2025 - The Comic Chronicles, scheduled to be held from August 25th to 30th at JIPMER, Puducherry. Your participation is vital to making this event a success, and we are excited to welcome you to a vibrant lineup of activities and discussions!\n\nRegistration Details:\n\n- Delegate Name: {reg.name}\n- College: {reg.college_name}\n- Tier: {reg.tier.upper()}\n- Delegate ID: {reg.user_id}\n- Date of Registration: {reg.created_at.strftime('%m/%d/%Y')}\n\nYou can use your Delegate ID {reg.user_id} to complete event registration through our official website.\n\nPlease carry a copy of this email and your college ID at the venue for smooth entry. Event guidelines and schedules will be shared soon.\n\nFor help, contact us at jsa.jipmer@gmail.com.\n\nWe look forward to hosting you at Spandan 2025!\n\nWarm regards,\nSuriya\nPresident, JIPMER Student Association"
    )
    return Response({"message": "Delegate card verified"}, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def reject_delegate_registration(request, pk):
    try:
        reg = DelegateCardRegistration.objects.get(pk=pk)
    except DelegateCardRegistration.DoesNotExist:
        return Response({'error': 'Registration not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.status = 'rejected'
    reg.is_verified = False
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="❌ Delegate Card Rejected – Spandan 2025",
        message=f"Hi {reg.name},\n\nYour delegate card registration (Tier: {reg.tier}) has been rejected.\nPlease check your submission and try again.\n\nFor any query, contact us at jsa.jipmer@gmail.com.\n\nRegards,\nTeam Spandan"
    )
    return Response({'message': 'Delegate card rejected'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_verified_delegate_cards(request):
    regs = DelegateCardRegistration.objects.filter(is_verified=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delegate Cards"
    headers = ["User ID", "Name", "Email", "Phone", "College", "Tier", "Amount", "Status", "Transaction ID", "Verified At", "Date"]
    ws.append(headers)
    for r in regs:
        ws.append([
            r.user_id, r.name, r.email, r.phone, r.college_name, r.tier,
            r.amount, r.status, r.transaction_id,
            r.verified_at.strftime("%Y-%m-%d %H:%M") if r.verified_at else "",
            r.created_at.strftime("%Y-%m-%d")
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=delegate_cards.xlsx'
    wb.save(response)
    return response

EVENT_PRICE_MAP = {
    # Fine Arts
    'face_painting': 100,
    'pot_painting': 100,
    'mehendi': 100,
    'sketching': 100,
    'painting': 100,

    # Dance
    'solo_dance': 150,
    'duet_dance': 150,
    'chorea_theme': 250,
    'chorea_nontheme': 250,
    'show_down': 150,

    # Music
    'tinnitus': 250,
    'alaap': 250,
    'euphony': 150,
    'raag_rangmanch': 150,
    'solo_singing': 150,
    'solo_instrumental': 150,

    # Drama
    'play': 125,
    'skit': 125,
    'mime': 125,
    'adzap': 125,
    'variety': 125,
    'dernier_cri': 250,

    # Sports
    'cricket': 300,
    'football': 300,
    'basketball_men': 300,
    'basketball_women': 300,
    'volleyball_men': 300,
    'volleyball_women': 300,
    'hockey_men': 300,
    'futsal_men': 300,
    'futsal_women': 300,
    'chess_bullet': 200,
    'chess_rapid': 200,
    'chess_blitz': 200,
    'chess_team' : 200,
    'carroms': 150,
    'throwball_men': 300,
    'throwball_women': 300,
    # Badminton (all subcategories same price)
    'badminton_men_singles': 300,
    'badminton_women_singles': 300,
    'badminton_men_doubles': 300,
    'badminton_women_doubles': 300,
    'badminton_mixed_doubles': 300,
    
    # Tennis
    'tennis_men_singles': 300,
    'tennis_women_singles': 300,
    'tennis_men_doubles': 300,
    'tennis_women_doubles': 300,
    'tennis_mixed_doubles': 300,
    
    # Table Tennis
    'table_tennis_men_singles': 300,
    'table_tennis_women_singles': 300,
    'table_tennis_men_doubles': 300,
    'table_tennis_women_doubles': 300,
    
    # Athletics
    'athletics_100m': 200, 'athletics_200m': 200, 'athletics_400m': 200,
    'athletics_800m': 200, 'athletics_1500m': 200, 'athletics_4x100m': 200,
    'athletics_shot_put': 200, 'athletics_discus_throw': 200, 'athletics_long_jump': 200,
    # ... all athletics subcategories 200 ...
    
    # Aquatics
    'aquatics_men_50m_freestyle': 200, 'aquatics_men_50m_backstroke': 200,
    'aquatics_men_50m_breaststroke': 200,
    'aquatics_men_50m_butterfly': 200,
    'aquatics_men_4x50m_freestyle_relay': 200,
    'aquatics_women_50m_freestyle': 200, 'aquatics_women_50m_backstroke': 200,
    'aquatics_women_50m_breaststroke': 200, 'aquatics_women_50m_butterfly': 200,
    'aquatics_mixed_4x50m_freestyle_relay': 200,

    # Literary
    'malarkey': 150,
    'shipwrecked': 150,
    'turncoat': 150,
    'scrabble': 150,
    'formal_debate': 250,
    'cryptic_crossword': 150,
    'ppt_karaoke': 150,
    'potpourri': 150,
    'india_quiz': 150,
    'fandom_quiz': 150,
    'sports_quiz': 150,
    'rewind_quiz': 150,
    'tj_jaishankar_memorial_quiz': 250,
    'jam': 250,
}

# Event Registration Views
class EventRegisterView(APIView):
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        regs = EventRegistration.objects.all()
        serializer = EventRegistrationSerializer(regs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def validate_delegate_ids(self, delegate_ids, email):
        if not delegate_ids:
            return True
        for delegate_id in delegate_ids:
            try:
                card = DelegateCardRegistration.objects.get(user_id=delegate_id, is_verified=True)
                if card.email != email:
                    return False
            except DelegateCardRegistration.DoesNotExist:
                return False
        return True

    def post(self, request):
        try:
            data = request.data.dict()  # Convert QueryDict to regular dict
            
            # Parse delegate_info if provided
            if 'delegate_info' in data and isinstance(data['delegate_info'], str):
                try:
                    data['delegate_info'] = json.loads(data['delegate_info'])
                except json.JSONDecodeError:
                    return Response(
                        {"delegate_info": "Invalid JSON format"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Parse events field
            if 'events' in data and isinstance(data['events'], str):
                try:
                    data['events'] = json.loads(data['events'])
                except json.JSONDecodeError:
                    return Response(
                        {"events": "Invalid JSON format"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            email = data.get("email", "").lower().strip()
            selected_events = data.get("events", [])
            
            if not isinstance(selected_events, list) or len(selected_events) == 0:
                return Response(
                    {"events": "At least one event must be selected"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            valid_events = (set().union(*CATEGORY_EVENT_MAP.values()) | BADMINTON_EVENTS | TENNIS_EVENTS | TABLE_TENNIS_EVENTS | ATHLETICS_EVENTS | AQUATICS_EVENTS)

            for event in selected_events:
                if event not in valid_events:
                    return Response(
                        {"error": f"Invalid event: {event}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Check event requirements
            needs_delegate = any(
                event not in EXEMPT_EVENTS
                and not (
                    (event in SPORTS_EVENTS or
                    event in BADMINTON_EVENTS or
                    event in TENNIS_EVENTS or
                    event in TABLE_TENNIS_EVENTS or
                    event in ATHLETICS_EVENTS or
                    event in AQUATICS_EVENTS) and
                    PassPurchase.objects.filter(email=email, pass_type='sports', is_verified=True).exists()
                ) or
                (event in CULT_EVENTS and PassPurchase.objects.filter(email=email, pass_type='cult', is_verified=True).exists()) or
                (event in LIT_EVENTS and PassPurchase.objects.filter(email=email, pass_type__in=['lit_lit', 'lit_premium'], is_verified=True).exists()) or
                (event in QUIZ_EVENTS and PassPurchase.objects.filter(email=email, pass_type__in=['lit_quiz', 'lit_premium'], is_verified=True).exists())
                for event in selected_events
            )

            if needs_delegate and not DelegateCardRegistration.objects.filter(email=email, is_verified=True).exists():
                return Response(
                    {"error": "Delegate card or valid pass required for selected events"},
                    status=status.HTTP_403_FORBIDDEN
                )



            amount = sum(EVENT_PRICE_MAP.get(event, 0) for event in selected_events)
            teammate_count = len([t for t in (data.get('delegate_info') or [])
                                    if t.get('delegate_id') and t.get('email')]) or 1
            data['amount'] = amount*teammate_count
            
            serializer = EventRegistrationSerializer(data=data)
            if serializer.is_valid():
                reg = serializer.save()
                
                # Auto-generate user_id if not set
                if not reg.user_id:
                    reg.user_id = f"USER-EV-{reg.id:05d}"
                    reg.save()
                
                return Response({
                    "message": "Event registration submitted",
                    "user_id": reg.user_id,
                    "registration": EventRegistrationSerializer(reg).data
                }, status=status.HTTP_201_CREATED)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {"error": "Server error", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# THIS IS THE CORRECTED FUNCTION
@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def verify_event_registration(request, pk):
    try:
        reg = EventRegistration.objects.get(pk=pk)
    except EventRegistration.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.is_verified = True
    reg.status = 'approved'
    reg.verified_at = timezone.now()
    reg.save()

    # Create the formatted string for the event list first.
    event_list_str = "\n".join([f"- {event}" for event in reg.events])

    send_smtp_email(
        to_email=reg.email,
        subject="✅ Event Registration Verified – Spandan 2025",
        # Insert the pre-formatted string into the main message.
        message=f"Dear {reg.name},\n\nWe are thrilled to confirm your registration for the following events at Spandan 2025 - The Comic Chronicles:\n{event_list_str}\n\nRegistration Details:\nName: {reg.name}\nCollege: {reg.college}\nEmail: {reg.email}\nTotal Paid: ₹{reg.amount}\nEvent ID: {reg.user_id}\n- Date: {reg.created_at.strftime('%m/%d/%Y')}\n\nPlease carry a copy of this confirmation email and your delegate ID (if applicable) during the event.\n\nIf you have questions or need help, feel free to write to us at jsa.jipmer@gmail.com.\n\nAll the best and see you soon at Spandan 2025!\n\nWarm regards,\nTeam Spandan"
    )
    return Response({"message": "Event registration verified"}, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def reject_event_registration_soft(request, pk):
    try:
        reg = EventRegistration.objects.get(pk=pk)
    except EventRegistration.DoesNotExist:
        return Response({'error': 'Registration not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.status = 'rejected'
    reg.is_verified = False
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="❌ Event Registration Rejected – Spandan 2025",
        message=f"Hi {reg.name},\n\nYour event registration has been rejected.\n\nEvents:\n{'\n'.join(f'- {e}' for e in reg.events)}\n\nPlease review your payment and try again if needed.\n\nFor any query, contact us at jsa.jipmer@gmail.com.\n\nRegards,\nTeam Spandan"
    )
    return Response({'message': 'Event registration rejected'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_verified_event_registrations(request):
    try:
        regs = EventRegistration.objects.filter(is_verified=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Event Registrations"
        
        # Headers
        headers = [
            "Type", "User ID", "Name", "Email", "Phone", "College",
            "Events", "Amount", "Status", "Transaction ID",
            "Verified At", "Date"
        ]
        ws.append(headers)
        
        for r in regs:
            # Main registrant
            ws.append([
                "Main Registrant",
                r.user_id, r.name, r.email, r.phone, r.college,
                ", ".join(r.events),
                r.amount, r.status, r.transaction_id,
                r.verified_at.strftime("%Y-%m-%d %H:%M") if r.verified_at else "",
                r.created_at.strftime("%Y-%m-%d")
            ])
            
            # Teammates
            for i, teammate in enumerate(r.delegate_info or [], 1):
                # Get delegate details from database
                delegate = DelegateCardRegistration.objects.filter(
                    user_id=teammate.get('delegate_id'),
                    email=teammate.get('email')
                ).first()
                
                ws.append([
                    f"Teammate {i}",
                    teammate.get('delegate_id', ''),
                    delegate.name if delegate else '',
                    teammate.get('email', ''),
                    teammate.get('phone', ''),
                    delegate.college_name if delegate else '',
                    "", "", "", "", "", ""  # Empty fields for alignment
                ])
            
            # Add empty row between registrations
            ws.append([])
        
        # Adjust column widths
        for i, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(20, len(header) + 2)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=event_registrations.xlsx'
        wb.save(response)
        return response
        
    except Exception as e:
        return Response(
            {"error": "Export failed", "details": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )



@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_event_by_name(request, event_name):
    try:
        event_name = event_name.lower().strip()
        logger.info(f"Starting export for event: {event_name}")

        # Enhanced event hierarchy mapping including all sub-events
        EVENT_HIERARCHY = {
            # Sports sub-events
            'tennis': {
                'tennis_men_singles', 'tennis_women_singles',
                'tennis_men_doubles', 'tennis_women_doubles',
                'tennis_mixed_doubles'
            },
            'aquatics': {
                'aquatics_men_50m_freestyle', 'aquatics_men_50m_backstroke',
                'aquatics_men_50m_breaststroke', 'aquatics_men_50m_butterfly',
                'aquatics_men_4x50m_freestyle_relay',
                'aquatics_women_50m_freestyle', 'aquatics_women_50m_backstroke',
                'aquatics_women_50m_breaststroke', 'aquatics_women_50m_butterfly',
                'aquatics_mixed_4x50m_freestyle_relay'
            },
            'badminton': {
                'badminton_men_singles', 'badminton_women_singles',
                'badminton_men_doubles', 'badminton_women_doubles',
                'badminton_mixed_doubles'
            },
            'table_tennis': {
                'table_tennis_men_singles', 'table_tennis_women_singles',
                'table_tennis_men_doubles', 'table_tennis_women_doubles'
            },
            'athletics': {
                'athletics_100m', 'athletics_200m', 'athletics_400m',
                'athletics_800m', 'athletics_1500m', 'athletics_4x100m',
                'athletics_shot_put', 'athletics_discus_throw', 'athletics_long_jump'
            },
            'futsal': {
                'futsal_men', 'futsal_women'
            },
            'basketball': {
                'basketball_men', 'basketball_women'
            },
            'volleyball': {
                'volleyball_men', 'volleyball_women'
            },
            'hockey': {
                'hockey_men', 'hockey_women'
            },
            'throwball': {
                'throwball_men', 'throwball_women'
            },
            'chess': {
                'chess_bullet', 'chess_rapid', 'chess_blitz', 'chess_team'
            },
            
            # Dance sub-events
            'chorea': {
                'chorea_theme', 'chorea_nontheme'
            },
            
            # Fine Arts (individual events)
            'face_painting': {'face_painting'},
            'pot_painting': {'pot_painting'},
            'mehendi': {'mehendi'},
            'sketching': {'sketching'},
            'painting': {'painting'},
            
            # Dance (individual events)
            'solo_dance': {'solo_dance'},
            'duet_dance': {'duet_dance'},
            'show_down': {'show_down'},
            
            # Music
            'tinnitus': {'tinnitus'},
            'alaap': {'alaap'},
            'euphony': {'euphony'},
            'raag_rangmanch': {'raag_rangmanch'},
            'solo_singing': {'solo_singing'},
            'solo_instrumental': {'solo_instrumental'},
            
            # Drama
            'play': {'play'},
            'skit': {'skit'},
            'mime': {'mime'},
            'adzap': {'adzap'},
            'variety': {'variety'},
            'dernier_cri': {'dernier_cri'},
            
            # Literary
            'malarkey': {'malarkey'},
            'shipwrecked': {'shipwrecked'},
            'turncoat': {'turncoat'},
            'scrabble': {'scrabble'},
            'formal_debate': {'formal_debate'},
            'cryptic_crossword': {'cryptic_crossword'},
            'ppt_karaoke': {'ppt_karaoke'},
            'potpourri': {'potpourri'},
            
            # Quiz
            'india_quiz': {'india_quiz'},
            'fandom_quiz': {'fandom_quiz'},
            'sports_quiz': {'sports_quiz'},
            'rewind_quiz': {'rewind_quiz'},
            'tj_jaishankar_memorial_quiz': {'tj_jaishankar_memorial_quiz'},
            'jam': {'jam'},
            
            # Other sports (individual)
            'cricket': {'cricket'},
            'football': {'football'},
            'carroms': {'carroms'}
        }

        # Get all verified registrations
        all_regs = EventRegistration.objects.filter(is_verified=True).only(
            'user_id', 'name', 'email', 'phone', 'college',
            'events', 'amount', 'status', 'transaction_id',
            'verified_at', 'created_at', 'delegate_info'
        )
        
        matching_regs = []
        
        for reg in all_regs:
            try:
                # Normalize events data
                current_events = []
                if isinstance(reg.events, list):
                    current_events = [str(e).lower().strip() for e in reg.events]
                elif isinstance(reg.events, str):
                    try:
                        parsed = json.loads(reg.events)
                        current_events = [str(e).lower().strip() for e in parsed] if isinstance(parsed, list) else [str(parsed).lower().strip()]
                    except (json.JSONDecodeError, TypeError):
                        current_events = [e.strip().lower() for e in reg.events.split(',')]

                # Check for direct match or hierarchical relationship
                if event_name in current_events:
                    matching_regs.append(reg)
                else:
                    # Check if it's a subcategory of a main event
                    for main_event, sub_events in EVENT_HIERARCHY.items():
                        if event_name in sub_events and main_event in current_events:
                            matching_regs.append(reg)
                            break

            except Exception as e:
                logger.error(f"Error processing registration {reg.id}: {str(e)}", exc_info=True)
                continue

        if not matching_regs:
            logger.warning(f"No registrations found for {event_name}")
            return Response(
                {"error": f"No registrations found for {event_name}"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Format sheet title (Excel has 31-char limit for sheet names)
        sheet_title = (event_name.replace("_", " ").title()[:31])
        ws.title = sheet_title

        # Enhanced headers
        headers = [
            "Type", "User ID", "Name", "Email", "Phone", "College",
            "All Registered Events", "Amount", "Status",
            "Transaction ID", "Verified At", "Registration Date"
        ]
        ws.append(headers)

        # Data rows
        for reg in matching_regs:
            try:
                # Format events for display
                events_display = ", ".join(reg.events) if isinstance(reg.events, list) else str(reg.events)

                # Main registrant row
                ws.append([
                    "Main Registrant",
                    reg.user_id or "",
                    reg.name or "",
                    reg.email or "",
                    reg.phone or "",
                    reg.college or "",
                    events_display,
                    reg.amount or 0,
                    reg.status or "",
                    reg.transaction_id or "",
                    reg.verified_at.strftime("%Y-%m-%d %H:%M") if reg.verified_at else "",
                    reg.created_at.strftime("%Y-%m-%d") if reg.created_at else ""
                ])

                # Teammates rows
                if reg.delegate_info:
                    teammates = json.loads(reg.delegate_info) if isinstance(reg.delegate_info, str) else reg.delegate_info
                    
                    for i, teammate in enumerate(teammates, 1):
                        delegate = None
                        if isinstance(teammate, dict) and teammate.get('delegate_id'):
                            try:
                                delegate = DelegateCardRegistration.objects.filter(
                                    user_id=teammate['delegate_id']
                                ).first()
                            except Exception as e:
                                logger.error(f"Error fetching delegate {teammate.get('delegate_id')}: {str(e)}")

                        ws.append([
                            f"Teammate {i}",
                            teammate.get('delegate_id', ''),
                            delegate.name if delegate else teammate.get('name', ''),
                            teammate.get('email', ''),
                            teammate.get('phone', ''),
                            delegate.college_name if delegate else teammate.get('college', ''),
                            "", "", "", "", "", ""
                        ])

                # Add empty row between registrations
                ws.append([])

            except Exception as e:
                logger.error(f"Error generating row for registration {reg.id}: {str(e)}", exc_info=True)
                continue

        # Adjust column widths
        for i, header in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max(20, len(header) + 2)

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            status=status.HTTP_200_OK
        )
        filename = f"{event_name}_registrations.xlsx".replace(" ", "_")
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Export failed completely: {str(e)}", exc_info=True)
        return Response(
            {"error": "Export processing failed", "details": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Pass Purchase Views
class PassPurchaseView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = PassPurchaseSerializer(data=request.data)
        if serializer.is_valid():
            reg = serializer.save()
            return Response({
                "message": "Pass purchase submitted",
                "purchase": PassPurchaseSerializer(reg).data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        self.check_permissions(request)
        regs = PassPurchase.objects.all()
        serializer = PassPurchaseSerializer(regs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def verify_pass(request, pk):
    try:
        reg = PassPurchase.objects.get(pk=pk)
    except PassPurchase.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.is_verified = True
    reg.status = 'approved'
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject=f"✅ {reg.get_pass_type_display()} Verified – Spandan 2025",
        message=f"Dear {reg.name},\n\nWe're excited to confirm that your {reg.pass_type} for Spandan 2025 - The Comic Chronicles has been successfully verified!\n\nPass Details:\n- Name: {reg.name}\n- College: {reg.college_name}\n- Pass: {reg.pass_type}\n- Pass ID: {reg.user_id}\n- Amount Paid: ₹{reg.amount}\n- Date of Purchase: {reg.created_at.strftime('%m/%d/%Y')}\n\nYour pass allows you to participate in eligible events under this category. Please carry this confirmation and your college ID for smooth verification at the venue.\n\nFor any support, feel free to reach us at jsa.jipmer@gmail.com.\n\nWarm regards,\nTeam Spandan"
    )
    return Response({"message": "Pass verified"}, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def reject_pass_soft(request, pk):
    try:
        reg = PassPurchase.objects.get(pk=pk)
    except PassPurchase.DoesNotExist:
        return Response({'error': 'Pass not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.status = 'rejected'
    reg.is_verified = False
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject=f"❌ {reg.get_pass_type_display()} Rejected – Spandan 2025",
        message=f"Hi {reg.name},\n\nYour pass purchase ({reg.get_pass_type_display()}) has been rejected.\nPlease check the screenshot or contact support.\n\nFor any query, contact us at jsa.jipmer@gmail.com.\n\nRegards,\nTeam Spandan"
    )
    return Response({'message': 'Pass rejected'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_verified_passes(request):
    regs = PassPurchase.objects.filter(is_verified=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pass Purchases"
    headers = ["User ID", "Name", "Email", "Phone", "College", "Pass Type", "Amount", "Status", "Transaction ID", "Verified At", "Date"]
    ws.append(headers)
    for r in regs:
        ws.append([
            r.user_id, r.name, r.email, r.phone, r.college_name,
            r.get_pass_type_display(), r.amount, r.status, r.transaction_id,
            r.verified_at.strftime("%Y-%m-%d %H:%M") if r.verified_at else "",
            r.created_at.strftime("%Y-%m-%d")
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pass_purchases.xlsx'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def check_delegate_eligibility(request, email):
    email = email.lower()
    try:
        reg = DelegateCardRegistration.objects.get(email=email)
        return Response({
            'is_verified': reg.is_verified,
            'user_id': reg.user_id,
            'tier': reg.tier
        })
    except DelegateCardRegistration.DoesNotExist:
        return Response({'is_verified': False})
    except DelegateCardRegistration.MultipleObjectsReturned:
        regs = DelegateCardRegistration.objects.filter(email=email)
        verified = any(reg.is_verified for reg in regs)
        return Response({
            'is_verified': verified,
            'user_id': [reg.user_id for reg in regs],
            'tier': [reg.tier for reg in regs]
        })
